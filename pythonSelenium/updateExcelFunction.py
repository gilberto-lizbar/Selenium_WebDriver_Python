import openpyxl


def update_excel_data(file_path, colName, searchTerm, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    dict1 = {}

    # Search the column price in Excel sheet
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            dict1["col"] = i  # store column number in dictionary
    print(dict1)
    # Search for Apple in the sheet to get the row where is located the price of apple
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                dict1["row"] = i
    print(dict1)

    actual_value = sheet.cell(row=dict1["row"], column=dict1["col"]).value
    sheet.cell(row=dict1["row"], column=dict1["col"]).value = new_value
    print(f"Value of column {dict1["col"]}, row {dict1["row"]} was modified from {actual_value} to "
          f"{new_value}")
    # Save excel book
    book.save(file_path)


update_excel_data("/Users/gilberto.barraza/Desktop/Repo_Cursos/Documents_repo/download.xlsx",
                  "color", "Orange", "Yellow")
