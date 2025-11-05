import openpyxl

book = openpyxl.load_workbook("/Users/gilberto.barraza/Desktop/Repo_Cursos"
                              "/Documents_repo/pythonDemo.xlsx")

sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Gilberto"   # Type in cell 2, column 2 in excel sheet 'Gilberto'

print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet['B2'].value)

# iterate cells in column 2
for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i, column=2).value)
# iterate all cells and rows
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)
# iterate rows and columns for 2 rows
for i in range(1, 3):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)
# iterate rows and columns if value is == "testcase3" then print row data
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase3":
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)
# iterate rows and columns if value is == "testcase3" then print row data
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase3":
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)

dict1 = {}
# iterate rows and columns if value is == "testcase3" then print row data and
# add the values of the row in a dictionary
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase3":
        for j in range(1, sheet.max_column + 1):
            dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            print(sheet.cell(row=i, column=j).value)
print(dict1)








