import openpyxl

file_path = ("/Users/gilberto.barraza/Desktop/Repo_Cursos"
             "/Documents_repo/download.xlsx")
book = openpyxl.load_workbook(file_path)
sheet = book.active
dict1 = {}

# Search the column price in Excel sheet
for i in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=i).value == "price":
        dict1["col"] = i  # store column number in dictionary
print(dict1)
# Search for Apple in the sheet to get the row where is located the price of apple
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        if sheet.cell(row=i, column=j).value == "Apple":
            dict1["row"] = i
            # Modify apple Price
            sheet.cell(row=dict1["row"], column=dict1["col"]).value = "500"
print(dict1)

# Print all the cell data from the row where Apple is locate
apple_row = 0
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        if sheet.cell(row=i, column=j).value == "Apple":
            apple_row = i
print(apple_row)
# Print all the cell data from the row where Apple is locate
print("******************")
for i in range(1, sheet.max_column + 1):
    print(sheet.cell(row=apple_row, column=i).value)
print("******************")
# Print apple price with the row and column coordinates
print("Apple price $", sheet.cell(row=dict1["row"], column=dict1["col"]).value)
# Modify apple Price
sheet.cell(row=dict1["row"], column=dict1["col"]).value = "345"
print("New Apple price $", sheet.cell(row=dict1["row"], column=dict1["col"]).value)
# Save excel book
book.save(file_path)
